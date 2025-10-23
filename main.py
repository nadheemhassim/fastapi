from fastapi import FastAPI, Query
from typing import Generator, List, Optional
import pandas as pd
import numpy as np
from openpyxl import load_workbook

app = FastAPI(title="Streaming Excel Filter API", version="1.3")

DATASET_1 = ""

def stream_excel_in_chunks(file_path: str, sheet_name: str = None, chunk_size: int = 5000) -> Generator[pd.DataFrame, None, None]:
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    chunk = []
    for row in rows:
        chunk.append(row)
        if len(chunk) >= chunk_size:
            yield pd.DataFrame(chunk, columns=headers)
            chunk = []
    if chunk:
        yield pd.DataFrame(chunk, columns=headers)

def df_to_json_safe(df: pd.DataFrame) -> List[dict]:
    df_clean = df.replace([float("inf"), float("-inf")], np.nan)
    result = df_clean.to_dict(orient="records")
    for row in result:
        for k, v in row.items():
            if isinstance(v, float) and (pd.isna(v) or np.isinf(v)):
                row[k] = None
    return result

def filter_chunk(
    df: pd.DataFrame,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    providers: Optional[List[str]] = None,
    facilities: Optional[List[str]] = None,
    states: Optional[List[str]] = None,
    payers: Optional[List[str]] = None,
    encounter_types: Optional[List[str]] = None
) -> pd.DataFrame:
    """Filter a chunk of Excel data by multiple criteria including encounter type."""

    if providers:
        df = df[df["Provider Name"].isin(providers)]
    if facilities:
        df = df[df["Facility Name"].isin(facilities)]
    if states:
        df = df[df["State"].isin(states)]
    if payers:
        df = df[df["Insurance Name"].isin(payers)]

    if encounter_types:
        encounter_frames = []

        if "Initial" in encounter_types:
            temp = df.copy()
            if start_date:
                temp = temp[pd.to_datetime(temp["Transaction Date"], errors="coerce") >= pd.to_datetime(start_date)]
            if end_date:
                temp = temp[pd.to_datetime(temp["Transaction Date"], errors="coerce") <= pd.to_datetime(end_date)]
            encounter_frames.append(temp)

        if "Follow-up" in encounter_types:
            temp = df.copy()
            if start_date:
                temp = temp[pd.to_datetime(temp["Date"], errors="coerce") >= pd.to_datetime(start_date)]
            if end_date:
                temp = temp[pd.to_datetime(temp["Date"], errors="coerce") <= pd.to_datetime(end_date)]
            encounter_frames.append(temp)

        if encounter_frames:
            df = pd.concat(encounter_frames).drop_duplicates()

    return df

@app.get("/read-dataset")
def read_dataset1(
    limit: int = 10,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    providers: Optional[List[str]] = Query(None, description="Multi-select providers"),
    facilities: Optional[List[str]] = Query(None, description="Multi-select facilities"),
    states: Optional[List[str]] = Query(None, description="Multi-select states"),
    payers: Optional[List[str]] = Query(None, description="Multi-select payers"),
    encounter_types: Optional[List[str]] = Query(None, description="Initial = Transaction Date, Follow-up = Date, or both")
):
    try:
        rows = []
        total_rows = 0

        for chunk in stream_excel_in_chunks(DATASET_1, chunk_size=5000):
            filtered = filter_chunk(
                chunk,
                start_date=start_date,
                end_date=end_date,
                providers=providers,
                facilities=facilities,
                states=states,
                payers=payers,
                encounter_types=encounter_types
            )
            if not filtered.empty:
                rows.append(filtered)
                total_rows += len(filtered)
            if total_rows >= limit:
                break

        if not rows:
            return {"message": "No data found for the given filters."}

        df = pd.concat(rows).head(limit)
        return df_to_json_safe(df)

    except Exception as e:
        return {"error": str(e)}
